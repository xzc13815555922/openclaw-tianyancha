#!/usr/bin/env python3
"""
天眼查数据查询报告模块
功能：根据多列规则输出统计报告 / 导出 Excel 企业清单

═══════════════════════════════════════════════════════════════
⚠️ 重要：所有 Excel 导出必须通过本脚本完成，    
          禁止在其他 Python 脚本中直接写 openpyxl 生成 Excel。
          外部调用本脚本时，请通过以下方式之一：
            subprocess.run(['python3', 'scripts/query.py', '--building', '金融城', '--output', 'excel'])
            或直接在本脚本的 main() 中调用 query() + export_excel()
═══════════════════════════════════════════════════════════════

使用方式：
    python3 scripts/query.py                         # 全量统计
    python3 scripts/query.py --region 盐南高新区      # 按区县筛选
    python3 scripts/query.py --building 金融城      # 按大楼宇筛选
    python3 scripts/query.py --street 黄海街道        # 按街道筛选
    python3 scripts/query.py --year 2026             # 按成立年份筛选
    python3 scripts/query.py --year 2026 --month 6   # 按成立年份+月份筛选
    python3 scripts/query.py --year 2026 --from-month 6  # 2026年6月起
    python3 scripts/query.py --not 个体              # 排除包含某关键词的记录
    python3 scripts/query.py --output excel          # 导出 Excel

Excel 导出规范（query_summary_pdf.py 定义的格式）：
    列序：区县 | 街道 | 大楼宇 | 小楼宇 | 注册时间 | 企业名称 | 联系号码 | 其余字段...
    排序：区县 → 街道 → 大楼宇 → 注册时间倒序（最新在前），小楼宇不参与排序但正常展示
    命名：按筛选条件自动命名，如「盐南高新区_黄海街道_2026年新注册企业清单.xlsx」

示例：
    # 盐南高新区 2026年新建企业清单（不含个体工商户）
    python3 scripts/query.py --region 盐南高新区 --year 2026 --not 个体 --output excel

    # 金融城全量注册企业清单
    python3 scripts/query.py --building 金融城 --output excel

    # 步凤镇 2026年新注册企业清单
    python3 scripts/query.py --street 步凤镇 --year 2026 --not 个体 --output excel

    # 华邦国际企业清单
    python3 scripts/query.py --building 华邦国际 --output excel

    # 国际创投中心企业清单
    python3 scripts/query.py --building 国际创投中心 --output excel
"""""

import sys
import sqlite3
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

DB_PATH = Path(__file__).parent.parent / "tianyancha.db"
OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── 固定列序（Excel 输出时前7列顺序）────────────────────────────────
EXCEL_COL_ORDER = [
    "region",           # 区县
    "street",           # 街道
    "building",         # 大楼宇
    "small_building",   # 小楼宇
    "establishment_date",  # 注册时间
    "enterprise_name",  # 企业名称
    "mobile_phone",     # 联系号码
]

# 其余字段（按数据库原有顺序追加，不含前7列已有的字段）
REMAINING_COLS = [
    "unified_social_credit_code",
    "registration_status",
    "legal_representative",
    "registered_capital",
    "province",
    "city",
    "county",
    "registered_address",
    "more_phones",
    "email",
    "enterprise_type",
    "taxpayer_id",
    "registration_number",
    "organization_code",
    "social_security_count",
    "social_security_year",
    "business_term",
    "industry_sector",
    "industry_major",
    "industry_mid",
    "industry_minor",
    "qcc_sector",
    "qcc_major",
    "qcc_mid",
    "qcc_minor",
    "enterprise_scale",
    "former_name",
    "english_name",
    "website",
    "contact_address",
    "contact_address_zip",
    "company_intro",
    "business_scope",
    "account_manager",
    "collected_date",
    "source_file",
    "created_at",
]

COLUMN_DISPLAY_NAMES = {
    "region": "区县",
    "street": "街道",
    "building": "大楼宇",
    "small_building": "小楼宇",
    "establishment_date": "注册时间",
    "enterprise_name": "企业名称",
    "mobile_phone": "联系号码",
    "unified_social_credit_code": "统一社会信用代码",
    "registration_status": "登记状态",
    "legal_representative": "法定代表人",
    "registered_capital": "注册资本",
    "province": "省份",
    "city": "城市",
    "county": "区县（原始）",
    "registered_address": "注册地址",
    "more_phones": "更多电话",
    "email": "邮箱",
    "enterprise_type": "企业(机构)类型",
    "taxpayer_id": "纳税人识别号",
    "registration_number": "注册号",
    "organization_code": "组织机构代码",
    "social_security_count": "参保人数",
    "social_security_year": "参保人数所属年报",
    "business_term": "营业期限",
    "industry_sector": "国标行业门类",
    "industry_major": "国标行业大类",
    "industry_mid": "国标行业中类",
    "industry_minor": "国标行业小类",
    "qcc_sector": "企查查行业门类",
    "qcc_major": "企查查行业大类",
    "qcc_mid": "企查查行业中类",
    "qcc_minor": "企查查行业小类",
    "enterprise_scale": "企业规模",
    "former_name": "曾用名",
    "english_name": "英文名",
    "website": "网址",
    "contact_address": "通信地址",
    "contact_address_zip": "通信地址邮编",
    "company_intro": "企业简介",
    "business_scope": "经营范围",
    "account_manager": "客户经理",
    "collected_date": "采集日期",
    "source_file": "来源文件",
    "created_at": "入库时间",
}

ALLOWED_GROUP_BY = {"region", "street", "building", "small_building", "enterprise_type", "registration_status"}

# ── WHERE 构建 ────────────────────────────────────────────────────
def build_where_clause(filters):
    conditions = []
    params = []
    # 预扫 year_eq 以便 month_gte 拼接 YYYY-MM
    year_val = None
    month_gte_val = None
    for field, op, value in filters:
        if op == "year_eq":
            year_val = value
        elif op == "month_gte":
            month_gte_val = value
    
    for field, op, value in filters:
        if op == "=":
            conditions.append(f"{field} = ?")
            params.append(value)
        elif op == "like":
            conditions.append(f"{field} LIKE ?")
            params.append(f"%{value}%")
        elif op == "not_like":
            conditions.append(f"({field} IS NULL OR {field} NOT LIKE ?)")
            params.append(f"%{value}%")
        elif op == "year_eq":
            conditions.append(f"SUBSTR({field}, 1, 4) = ?")
            params.append(str(value))
        elif op == "month_eq":
            month_str = str(value).zfill(2)
            conditions.append(f"SUBSTR({field}, 6, 2) = ?")
            params.append(month_str)
        elif op == "month_gte":
            prefix = year_val if year_val else ""
            if prefix:
                ym = f"{prefix}-{str(value).zfill(2)}"
                conditions.append(f"SUBSTR({field}, 1, 7) >= ?")
                params.append(ym)
            else:
                conditions.append(f"SUBSTR({field}, 6, 2) >= ?")
                params.append(str(value).zfill(2))
        elif op == "!=":
            conditions.append(f"{field} != ?")
            params.append(value)
    return conditions, params

# ── 查询 ──────────────────────────────────────────────────────────
def query(filters=None, group_by=None, limit=None, is_company=False, exclude_keyword=None):
    if group_by and group_by not in ALLOWED_GROUP_BY:
        raise ValueError(f"不支持的分组字段: {group_by}，允许: {sorted(ALLOWED_GROUP_BY)}")

    conn = sqlite3.connect(str(DB_PATH))
    try:
        where_parts, params = build_where_clause(filters) if filters else ([], [])

        if is_company:
            where_parts.append("(enterprise_name NOT LIKE '%个体%' AND enterprise_type NOT LIKE '%个体%')")
        if exclude_keyword:
            where_parts.append("(enterprise_name NOT LIKE ? AND (enterprise_type IS NULL OR enterprise_type NOT LIKE ?))")
            params.extend([f"%{exclude_keyword}%", f"%{exclude_keyword}%"])

        where_sql = "WHERE " + " AND ".join(where_parts) if where_parts else ""

        if group_by:
            sql = f"SELECT '{group_by}', {group_by}, COUNT(*) FROM enterprise_detail {where_sql} GROUP BY {group_by} ORDER BY COUNT(*) DESC"
            rows = conn.execute(sql, params).fetchall()
            total = conn.execute(f"SELECT COUNT(*) FROM enterprise_detail {where_sql}", params).fetchone()[0]
            return {"type": "group", "group_by": group_by, "rows": rows, "total": total}
        else:
            all_cols = EXCEL_COL_ORDER + REMAINING_COLS
            cols_sql = ", ".join(all_cols)
            sql = f"""
                SELECT {cols_sql}
                FROM enterprise_detail
                {where_sql}
                ORDER BY
                    region,
                    street,
                    building,
                    CASE WHEN length(establishment_date) >= 10 THEN 1 ELSE 0 END DESC,
                    substr(establishment_date, 1, 10) DESC,
                    enterprise_name
            """
            if limit:
                sql += f" LIMIT {limit}"
            rows = conn.execute(sql, params).fetchall()
            total = conn.execute(f"SELECT COUNT(*) FROM enterprise_detail {where_sql}", params).fetchone()[0]
            return {"type": "list", "rows": rows, "total": total, "columns": all_cols}
    finally:
        conn.close()

# ── 控制台输出 ────────────────────────────────────────────────────
def format_results(result):
    if result["type"] == "group":
        print(f"\n{'='*60}")
        print(f"  {result['group_by']} 分布统计 (共 {result['total']} 条)")
        print(f"{'='*60}")
        print(f"{'名称':<30} {'数量':>10}")
        print("-" * 40)
        for row in result["rows"]:
            name = row[1] if row[1] else "(未标注)"
            print(f"{name:<30} {row[2]:>10}")
        print("-" * 40)
        print(f"合计: {result['total']}")
    else:
        print(f"\n{'='*60}")
        print(f"  查询结果 (共 {result['total']} 条)")
        print(f"{'='*60}")
        headers = result["columns"]
        print(" | ".join(f"{COLUMN_DISPLAY_NAMES.get(h, h):<20}" for h in headers[:7]))
        print("-" * 140)
        for row in result["rows"][:30]:
            vals = [str(v) if v else "" for v in row[:7]]
            print(" | ".join(v[:20].ljust(20) for v in vals))
        if len(result["rows"]) > 30:
            print(f"... (还有 {len(result['rows']) - 30} 条，已截断)")
        print("-" * 140)
        print(f"合计: {result['total']}")

# ── Excel 导出 ───────────────────────────────────────────────────
def export_excel(result, filters):
    if result["type"] != "list":
        print("❌ Excel 导出仅支持详细列表查询")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业清单"

    # 生成文件名
    filename = build_filename(filters, result["total"])
    output_path = OUTPUT_DIR / filename

    # 列名
    display_headers = [COLUMN_DISPLAY_NAMES.get(c, c) for c in result["columns"]]
    for col, h in enumerate(display_headers, 1):
        cell = ws.cell(1, col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for row_idx, row in enumerate(result["rows"], 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = val if val is not None else ""
            # 手机号特殊处理，防科学技术法
            col_name = result["columns"][col_idx - 1]
            if col_name == "mobile_phone" and val:
                cell.value = str(val)
                cell.data_type = 's'

    # 列宽
    col_widths = {
        1: 12,   # 区县
        2: 12,   # 街道
        3: 18,   # 大楼宇
        4: 20,   # 小楼宇
        5: 14,   # 注册时间
        6: 30,   # 企业名称
        7: 16,   # 联系号码
    }
    from openpyxl.utils import get_column_letter
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for col_idx in range(8, len(result["columns"]) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(str(output_path))
    print(f"\n✅ 已导出: {output_path}")
    print(f"   文件名: {filename}")
    print(f"   共 {result['total']} 条记录 | {len(result['columns'])} 列")

def build_filename(filters, total):
    """根据筛选条件生成文件名"""
    parts = []
    year_val = None
    month_val = None
    from_month_val = None
    street_val = None
    region_val = None
    building_val = None
    small_building_val = None
    exclude个体 = False

    for field, op, value in filters:
        if op == "year_eq":
            year_val = value
        elif op == "month_eq":
            month_val = value
        elif op == "month_gte":
            from_month_val = value
        elif field == "street":
            street_val = value
        elif field == "region":
            region_val = value
        elif field == "building":
            building_val = value
        elif field == "small_building":
            small_building_val = value
        elif field == "enterprise_name" and op == "not_like" and value == "个体":
            exclude个体 = True

    if region_val:
        parts.append(region_val)
    if street_val:
        parts.append(street_val)
    if building_val:
        parts.append(building_val)
    if year_val:
        if from_month_val:
            parts.append(f"{year_val}年{from_month_val}月起新注册")
        elif month_val:
            parts.append(f"{year_val}年{month_val}月新注册")
        else:
            parts.append(f"{year_val}年新注册")
    elif from_month_val:
        parts.append(f"{from_month_val}月起")
    if exclude个体:
        parts.append("非个体")

    base = "_".join(parts) if parts else "全量企业清单"
    return f"{base}_{total}条.xlsx"

# ── 主入口 ────────────────────────────────────────────────────────
def _parse_int(raw, name, lo=None, hi=None):
    """严格整数解析 + 范围检查 (友好报错退出)"""
    try:
        v = int(raw)
    except (TypeError, ValueError):
        print(f"❌ {name} 参数必须是整数，收到: {raw!r}", file=sys.stderr)
        sys.exit(2)
    if lo is not None and v < lo:
        print(f"❌ {name} 参数必须 ≥ {lo}，收到: {v}", file=sys.stderr)
        sys.exit(2)
    if hi is not None and v > hi:
        print(f"❌ {name} 参数必须 ≤ {hi}，收到: {v}", file=sys.stderr)
        sys.exit(2)
    return v


def main():
    filters = []
    group_by = None
    limit = None
    output = None
    is_company = False
    exclude_keyword = None

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        arg = args[i]
        if arg == "--region":
            filters.append(("region", "=", args[i+1])); i += 2
        elif arg == "--street":
            filters.append(("street", "=", args[i+1])); i += 2
        elif arg == "--building":
            filters.append(("building", "=", args[i+1])); i += 2
        elif arg == "--small_building":
            filters.append(("small_building", "=", args[i+1])); i += 2
        elif arg == "--year":
            year_v = _parse_int(args[i+1], "--year", lo=1900, hi=2100)
            filters.append(("establishment_date", "year_eq", year_v)); i += 2
        elif arg == "--month":
            # 单月筛选：月份 1-12 (不带 --year 也可用：仅按月份不按年筛选)
            month_v = _parse_int(args[i+1], "--month", lo=1, hi=12)
            filters.append(("establishment_date", "month_eq", month_v)); i += 2
        elif arg == "--from-month":
            # 起月筛选：月份 1-12 (与 --year 组合生成 YYYY-MM-01 范围下界)
            from_month_v = _parse_int(args[i+1], "--from-month", lo=1, hi=12)
            filters.append(("establishment_date", "month_gte", from_month_v)); i += 2
        elif arg == "--not":
            exclude_keyword = args[i+1]
            filters.append(("enterprise_name", "not_like", args[i+1])); i += 2
        elif arg == "--group":
            group_by = args[i+1]; i += 2
        elif arg == "--limit":
            limit = _parse_int(args[i+1], "--limit", lo=1); i += 2
        elif arg == "--output":
            output = args[i+1]; i += 2
        else:
            # P3 决策包反馈：未知参数不再静默忽略，直接报错退出
            print(f"❌ 未识别的参数: {arg}", file=sys.stderr)
            print(f"   已识别参数: --region --street --building --small_building", file=sys.stderr)
            print(f"                  --year --month --from-month --not --group --limit --output", file=sys.stderr)
            sys.exit(2)

    result = query(filters=filters, group_by=group_by, limit=limit, exclude_keyword=exclude_keyword)

    if result["type"] == "group":
        format_results(result)
    else:
        format_results(result)
        if output == "excel":
            export_excel(result, filters)

if __name__ == "__main__":
    main()